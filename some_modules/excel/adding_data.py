from openpyxl.worksheet.worksheet import Worksheet


def adding_data(worksheet: Worksheet) -> None:
    # Creating the header
    worksheet.cell(1, 1, 'Name')
    worksheet.cell(1, 2, 'Age')
    worksheet.cell(1, 3, 'Grade')


    students = [
        # name      age   grade
        ['João',    14,   5.5],
        ['Maria',   13,   9.7],
        ['Luiz',    15,   8.8],
        ['Alberto', 16,   10],
    ]

    # Adding data
    
    #* 1)
    
    # for i, student_row in enumerate(students, start= 2):
    #     for j, student_data in enumerate(students, start= 1):
    #         worksheet.cell(i, j, student_data)
        
    #* 2)

    for student in students: worksheet.append(student)
    # In this method,the rows are filled just after the last row.