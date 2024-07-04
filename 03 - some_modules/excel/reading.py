from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

def reading(worksheet: Worksheet) -> None:
    row: tuple[Cell]
    
    # for row in worksheet.iter_rows(max_row= 10):
    # for row in worksheet.iter_rows(min_row= 2):
    for row in worksheet.iter_rows():
        for cell in row:
            print(cell.value, end='\t')
        print()
        
    
    print(worksheet['B3'].value)