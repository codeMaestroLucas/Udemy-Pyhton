from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Union

def create_sheet() -> Union[Worksheet, Workbook]:
    
    workbook: Workbook = Workbook()
    
    sheet_name: str = 'Controll-grades'
    workbook.create_sheet(sheet_name, 0) # Creating the sheet and setting it up
    worksheet: Worksheet = workbook[sheet_name] # Selecting the sheet from the
    # available sheets


    #! It's important to SAVE the workbook after doing everything.
    return worksheet, workbook